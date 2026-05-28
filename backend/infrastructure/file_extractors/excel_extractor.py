import io

import pandas as pd
from openpyxl import load_workbook

from backend.domain.interfaces.file_extractor import FileExtractor
from backend.infrastructure.file_extractors.base_extractor import BaseExtractor


class ExcelExtractor(BaseExtractor, FileExtractor):
    def __init__(self, extension: str):
        self.extension = extension

    def extract(self, filename: str, file_bytes: bytes):
        parts = []
        tables = []

        if self.extension == "xlsx":
            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                parts.append(f"\n--- EXCEL SHEET: {sheet_name} ---")

                sheet_rows = []

                for row in sheet.iter_rows(values_only=True):
                    cells = [self._clean_text(cell) for cell in row]

                    if any(cells):
                        sheet_rows.append(cells)
                        parts.append(" | ".join(cells))

                tables.append(
                    {
                        "name": sheet_name,
                        "rows": sheet_rows,
                    }
                )

        else:
            excel_file = pd.ExcelFile(io.BytesIO(file_bytes), engine="xlrd")

            for sheet_name in excel_file.sheet_names:
                dataframe = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    header=None,
                )

                parts.append(f"\n--- EXCEL SHEET: {sheet_name} ---")

                sheet_rows = []

                for _, row in dataframe.iterrows():
                    cells = [self._clean_text(cell) for cell in row.tolist()]

                    if any(cells):
                        sheet_rows.append(cells)
                        parts.append(" | ".join(cells))

                tables.append(
                    {
                        "name": sheet_name,
                        "rows": sheet_rows,
                    }
                )

        return self._text_result(
            filename=filename,
            extension=self.extension,
            text="\n".join(parts),
            extractor_name=f"{self.extension}_tables_strategy",
            tables=tables,
        )
